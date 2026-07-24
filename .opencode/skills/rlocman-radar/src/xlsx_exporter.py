import json
import logging
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

logger = logging.getLogger(__name__)

SEARCH_COLUMNS = [
    "di", "score", "is_interesting", "is_read", "date_published",
    "article_url", "summary_ru", "title_ru", "author",
    "categories", "components",
]

SEARCH_HEADER_NAMES = {
    "di": "ID",
    "score": "Score",
    "is_interesting": "I",
    "is_read": "R",
    "date_published": "Date",
    "article_url": "URL",
    "summary_ru": "Summary",
    "title_ru": "Title",
    "author": "Author",
    "categories": "Cat",
    "components": "Cmp",
}

SEARCH_EDITABLE = {"is_interesting", "is_read"}

EDITABLE_HEADERS = {"I", "R"}
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
EDITABLE_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")


def export_to_xlsx(articles: list[dict], output_path: str,
                   columns=None, header_names=None, editable=None) -> str:
    if columns is None:
        columns = SEARCH_COLUMNS
    if header_names is None:
        header_names = SEARCH_HEADER_NAMES
    if editable is None:
        editable = SEARCH_EDITABLE

    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Articles"

    header = [header_names[c] for c in columns]
    for col_idx, col_name in enumerate(header, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = THIN_BORDER

    for row_idx, art in enumerate(articles, 2):
        for col_idx, col_name in enumerate(columns, 1):
            val = art.get(col_name, "")
            if col_name in ("is_interesting", "is_read"):
                val = "Y" if val else ""
            elif col_name == "article_url":
                url = val or f"https://www.rlocman.ru/shem/schematics.html?di={art.get('di', '')}"
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = f'=HYPERLINK("{url}", "{url}")'
                cell.border = THIN_BORDER
                continue
            elif val is None:
                val = ""

            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = THIN_BORDER
            if col_name in editable:
                cell.fill = EDITABLE_FILL
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_name == "summary_ru":
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    col_widths = {
        "di": 7, "score": 6, "is_interesting": 3, "is_read": 3,
        "date_published": 12, "article_url": 5, "summary_ru": 90,
        "title_ru": 55, "author": 18, "categories": 30, "components": 25,
    }
    for col_idx, col_name in enumerate(columns, 1):
        w = col_widths.get(col_name, 10)
        ws.column_dimensions[chr(64 + col_idx)].width = w

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"

    wb.save(str(path))
    logger.info("Exported %d articles to %s", len(articles), path)
    return str(path)


def import_from_xlsx(filepath: str, db) -> dict:
    path = Path(filepath)
    if not path.exists():
        raise FileNotFoundError(f"File not found: {filepath}")

    wb = load_workbook(str(path), data_only=True)
    ws = wb.active
    header = [cell.value for cell in ws[1]]
    col_map = {name: idx for idx, name in enumerate(header)}

    for col in EDITABLE_HEADERS:
        if col not in col_map:
            raise ValueError(f"Required column '{col}' not found in {filepath}")
    if "ID" not in col_map:
        raise ValueError("Required column 'ID' not found")

    updates_interesting = []
    updates_read = []
    rows = list(ws.iter_rows(min_row=2, values_only=True))

    for row in rows:
        pid = row[col_map["ID"]]
        if pid is None:
            continue
        current = db.get_article(pid)
        if current is None:
            logger.warning("Article di=%s not found in DB, skipping", pid)
            continue
        xlsx_i = bool(row[col_map["I"]])
        xlsx_r = bool(row[col_map["R"]])
        if xlsx_i != bool(current.get("is_interesting")):
            updates_interesting.append(pid)
        if xlsx_r != bool(current.get("is_read")):
            updates_read.append(pid)

    if updates_interesting:
        db.mark_interesting(updates_interesting)
    if updates_read:
        db.mark_read(updates_read)

    result = {
        "total_rows": len(rows),
        "updated_interesting": len(updates_interesting),
        "updated_read": len(updates_read),
    }
    logger.info("Import from %s: %d rows, %d I, %d R",
                filepath, result["total_rows"],
                result["updated_interesting"], result["updated_read"])
    return result