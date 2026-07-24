import logging
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

logger = logging.getLogger(__name__)

SEARCH_COLUMNS = [
    "id", "score",
    "is_interesting", "is_read",
    "date", "url", "summary_ru",
    "author", "title", "category", "tags",
    "comment",
]

SEARCH_HEADER_NAMES = {
    "id": "id",
    "score": "Score",
    "is_interesting": "I",
    "is_read": "R",
    "date": "Date",
    "url": "URL",
    "summary_ru": "Summary",
    "author": "Author",
    "title": "Title",
    "category": "Category",
    "tags": "Tags",
    "comment": "Reason",
}

SEARCH_EDITABLE = {
    "is_interesting": True,
    "is_read": True,
}

SEARCH_WIDTHS = {
    "id": 7, "score": 6, "is_interesting": 3, "is_read": 3,
    "date": 10, "url": 5, "summary_ru": 90,
    "author": 18, "title": 55, "category": 15, "tags": 30, "comment": 60,
}

WRAP_COLUMNS = {"summary_ru", "comment"}

THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
EDITABLE_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")


def _flatten_tags(tags) -> str:
    if isinstance(tags, list):
        return ", ".join(str(t) for t in tags)
    return str(tags) if tags else ""


def _col_letter(idx: int) -> str:
    return chr(64 + idx)


def export_to_xlsx(articles: list[dict], output_path: str,
                   columns=None, header_names=None, editable=None) -> str:
    if columns is None:
        columns = SEARCH_COLUMNS
    if header_names is None:
        header_names = SEARCH_HEADER_NAMES
    if editable is None:
        editable = SEARCH_EDITABLE

    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Articles"

    header = [header_names.get(c, c) for c in columns]
    for col_idx, col_name in enumerate(header, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = THIN_BORDER

    for row_idx, article in enumerate(articles, 2):
        for col_idx, col_name in enumerate(columns, 1):
            val = article.get(col_name, "")

            if col_name == "tags":
                val = _flatten_tags(val)
            elif col_name in ("is_interesting", "is_read"):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = "Y" if val else ""
                cell.border = THIN_BORDER
                cell.alignment = Alignment(horizontal="center", vertical="center")
                if editable.get(col_name):
                    cell.fill = EDITABLE_FILL
                continue
            elif col_name == "url" and val:
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = f'=HYPERLINK("{val}", "{val}")'
                cell.border = THIN_BORDER
                cell.font = Font(color="0563C1", underline="single")
                continue
            elif val is None:
                val = ""
            elif col_name == "score":
                val = int(val) if val else 0

            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = THIN_BORDER

            if col_name in WRAP_COLUMNS and val:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    for col_idx, col_name in enumerate(columns, 1):
        w = SEARCH_WIDTHS.get(col_name, 10)
        col_letter = _col_letter(col_idx)
        ws.column_dimensions[col_letter].width = w

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"

    wb.save(str(path))
    logger.info("Exported %d articles to %s", len(articles), path)
    return str(path)


def import_from_xlsx(db, xlsx_path: str) -> int:
    path = Path(xlsx_path)
    if not path.exists():
        raise FileNotFoundError(f"File not found: {xlsx_path}")

    wb = load_workbook(str(path), data_only=True)
    ws = wb.active

    header_row = [cell.value for cell in ws[1]]
    if not header_row:
        logger.warning("Empty header row in %s", xlsx_path)
        return 0

    try:
        id_idx = header_row.index("id")
    except ValueError:
        logger.error("No 'id' column found in %s", xlsx_path)
        return 0

    i_idx = header_row.index("I") if "I" in header_row else None
    r_idx = header_row.index("R") if "R" in header_row else None

    if i_idx is None and r_idx is None:
        logger.info("No I/R columns found, nothing to import")
        return 0

    updated = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        article_id = row[id_idx]
        if article_id is None:
            continue
        is_interesting = False
        is_read = False
        if i_idx is not None and row[i_idx] is not None:
            is_interesting = str(row[i_idx]).strip().upper() == "Y"
        if r_idx is not None and row[r_idx] is not None:
            is_read = str(row[r_idx]).strip().upper() == "Y"

        db.save_interesting(int(article_id), is_interesting, is_read)
        updated += 1

    logger.info("Imported I/R flags for %d articles from %s", updated, xlsx_path)
    return updated
