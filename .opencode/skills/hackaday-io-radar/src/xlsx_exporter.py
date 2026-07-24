import json
import logging
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

logger = logging.getLogger(__name__)

BASE_COLUMNS = [
    "id", "is_interesting", "is_read", "owner_name", "created_at",
    "url", "tags", "view_count", "followers_count", "summary_ru",
]

BASE_HEADER_NAMES = {
    "id": "id",
    "is_interesting": "I",
    "is_read": "R",
    "owner_name": "Owner",
    "created_at": "Created",
    "url": "URL",
    "tags": "Tags",
    "view_count": "Views",
    "followers_count": "Followers",
    "summary_ru": "Summary",
}

BASE_EDITABLE = {"is_interesting", "is_read"}

SEARCH_COLUMNS = [
    "id", "score", "is_interesting", "is_read", "created_at",
    "url", "summary_ru", "owner_name", "name",
    "view_count", "followers_count", "tags",
]

SEARCH_HEADER_NAMES = {
    "id": "id",
    "score": "Score",
    "is_interesting": "I",
    "is_read": "R",
    "created_at": "Date",
    "url": "URL",
    "summary_ru": "Summary",
    "owner_name": "Author",
    "name": "Title",
    "view_count": "Views",
    "followers_count": "Followers",
    "tags": "Tags",
}

SEARCH_EDITABLE = {"is_interesting", "is_read"}

EDITABLE_HEADERS = {"I", "R"}
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
EDITABLE_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")


def _flatten_tags(tags) -> str:
    if isinstance(tags, str):
        try:
            parsed = json.loads(tags)
            if isinstance(parsed, list):
                return ", ".join(parsed)
        except (json.JSONDecodeError, TypeError):
            pass
        return tags
    if isinstance(tags, list):
        return ", ".join(tags)
    return str(tags) if tags else ""


def export_to_xlsx(projects: list[dict], output_path: str,
                   columns=None, header_names=None, editable=None) -> str:
    if columns is None:
        columns = BASE_COLUMNS
    if header_names is None:
        header_names = BASE_HEADER_NAMES
    if editable is None:
        editable = BASE_EDITABLE

    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Projects"

    header = [header_names[c] for c in columns]
    for col_idx, col_name in enumerate(header, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = THIN_BORDER

    for row_idx, proj in enumerate(projects, 2):
        for col_idx, col_name in enumerate(columns, 1):
            val = proj.get(col_name, "")
            if col_name == "tags":
                val = _flatten_tags(val)
            elif col_name in ("is_interesting", "is_read"):
                val = "Y" if val else ""
            elif col_name == "url":
                slug = proj.get("slug") or proj.get("id", "")
                url = f"https://hackaday.io/project/{slug}"
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = f'=HYPERLINK("{url}", "{url}")'
                cell.border = THIN_BORDER
                continue
            elif val is None:
                val = ""

            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = THIN_BORDER
            if col_name in editable:
                cell.fill = EDITABLE_FILL
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_name == "summary_ru":
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    col_widths = {
        "id": 7, "score": 6, "is_interesting": 3, "is_read": 3,
        "owner_name": 18, "created_at": 12, "url": 5, "name": 55,
        "tags": 30, "view_count": 7, "followers_count": 10, "summary_ru": 90,
    }
    for col_idx, col_name in enumerate(columns, 1):
        w = col_widths.get(col_name, 10)
        ws.column_dimensions[chr(64 + col_idx)].width = w

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"

    wb.save(str(path))
    logger.info("Exported %d projects to %s", len(projects), path)
    return str(path)


def import_from_xlsx(filepath: str, db) -> dict:
    path = Path(filepath)
    if not path.exists():
        raise FileNotFoundError(f"File not found: {filepath}")

    wb = load_workbook(str(path), data_only=True)
    ws = wb.active

    header = [cell.value for cell in ws[1]]
    col_map = {name: idx for idx, name in enumerate(header)}

    for col in EDITABLE_HEADERS:
        if col not in col_map:
            raise ValueError(f"Required column '{col}' not found in {filepath}")
    if "id" not in col_map:
        raise ValueError("Required column 'id' not found")

    updates_interesting = []
    updates_read = []
    rows = list(ws.iter_rows(min_row=2, values_only=True))

    for row in rows:
        pid = row[col_map["id"]]
        if pid is None:
            continue
        current = db.get_project(pid)
        if current is None:
            logger.warning("Project id=%s not found in DB, skipping", pid)
            continue
        xlsx_i = bool(row[col_map["I"]])
        xlsx_r = bool(row[col_map["R"]])
        if xlsx_i != bool(current.get("is_interesting")):
            updates_interesting.append(pid)
        if xlsx_r != bool(current.get("is_read")):
            updates_read.append(pid)

    if updates_interesting:
        db.mark_interesting(updates_interesting)
    if updates_read:
        db.mark_read(updates_read)

    result = {
        "total_rows": len(rows),
        "updated_interesting": len(updates_interesting),
        "updated_read": len(updates_read),
    }
    logger.info("Import from %s: %d rows, %d is_interesting, %d is_read",
                filepath, result["total_rows"],
                result["updated_interesting"], result["updated_read"])
    return result