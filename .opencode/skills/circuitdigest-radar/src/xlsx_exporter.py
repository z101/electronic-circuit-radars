import json
import logging
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

logger = logging.getLogger(__name__)

COLUMNS = [
    "id", "score", "date", "url",
    "author", "title", "tags", "summary_ru",
]

HEADER_NAMES = {
    "id": "id",
    "score": "Score",
    "date": "Date",
    "url": "URL",
    "author": "Author",
    "title": "Title",
    "tags": "Tags",
    "summary_ru": "Summary",
}

SEARCH_COLUMNS = COLUMNS
SEARCH_HEADER_NAMES = HEADER_NAMES

THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)


def _flatten_tags(tags) -> str:
    if isinstance(tags, list):
        return ", ".join(tags)
    return str(tags) if tags else ""


def export_to_xlsx(articles: list[dict], output_path: str) -> str:
    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Articles"

    header = [HEADER_NAMES[c] for c in COLUMNS]
    for col_idx, col_name in enumerate(header, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = THIN_BORDER

    for row_idx, article in enumerate(articles, 2):
        for col_idx, col_name in enumerate(COLUMNS, 1):
            val = article.get(col_name, "")
            if col_name == "tags":
                val = _flatten_tags(val)
            elif col_name == "url" and val:
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = f'=HYPERLINK("{val}", "{val}")'
                cell.border = THIN_BORDER
                continue
            elif val is None:
                val = ""
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = THIN_BORDER

    col_widths = {"id": 7, "score": 6, "date": 10, "url": 5,
                  "author": 18, "title": 55, "tags": 30, "summary_ru": 90}
    for col_idx, col_name in enumerate(COLUMNS, 1):
        w = col_widths.get(col_name, 10)
        ws.column_dimensions[chr(64 + col_idx)].width = w

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"

    wb.save(str(path))
    logger.info("Exported %d articles to %s", len(articles), path)
    return str(path)